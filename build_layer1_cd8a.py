"""
Build layered feature vectors for CD8α IgV from:
  1) Excel diagram              → Layer 1 identity
  2) iCn3D interaction HTML     → Layer 4 H-bond density / partners

Inputs:
  - 1CD8_ig_diagram.xlsx              (sheet "1. IgV")
  - P01732-allinteraction.html   (iCn3D "All interactions" H-bond table)

Output:
  - CD8a-P01732.layers.json
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
EXCEL_DEFAULT = ROOT / "1CD8_ig_diagram.xlsx"
INTERACTION_DEFAULT = ROOT / "P01732-allinteraction.html"
OUT_DEFAULT = ROOT / "CD8a-P01732.layers.json"

CHAIN = "A"
PDB_ID = "1CD8"
UNIPROT_ID = "P01732"
DOMAIN = "IgV"

RES_PAT = re.compile(r"^([A-Z])(\d+)$")
HBOND_ROW_PAT = re.compile(
    r'<td align="center">([\d.]+)</td>'
    r'<td align="center"><button class="div0_selres" resid="([^"]+)\|([^"]+)">'
)
ATOM_PAT = re.compile(rf"([A-Z]{{3}}) \${UNIPROT_ID}\.A:(\d+)@([A-Za-z0-9]+)")

AA3_TO_1 = {
    "ALA": "A",
    "ARG": "R",
    "ASN": "N",
    "ASP": "D",
    "CYS": "C",
    "GLN": "Q",
    "GLU": "E",
    "GLY": "G",
    "HIS": "H",
    "ILE": "I",
    "LEU": "L",
    "LYS": "K",
    "MET": "M",
    "PHE": "F",
    "PRO": "P",
    "SER": "S",
    "THR": "T",
    "TRP": "W",
    "TYR": "Y",
    "VAL": "V",
}
BACKBONE_ATOMS = {"N", "O", "C", "CA"}

STRAND_FROM_IG = {
    1: "A",
    2: "B",
    3: "C",
    4: "C'",
    5: "C''",
    6: "D",
    7: "E",
    8: "F",
    9: "G",
}


def empty_layers() -> dict:
    return {
        "2_intrinsic": {},
        "3_structural": {},
        "4_interaction": {},
        "5_comparative": {},
    }


def segment_from_igstrand(ig: int) -> str:
    thousands = ig // 1000
    hundreds = (ig // 100) % 10
    if thousands == 1 and hundreds == 8:
        return "A'"
    return STRAND_FROM_IG.get(thousands, f"strand_{thousands}")


def is_anchor(ig: int) -> bool:
    return ig % 100 == 50


def weight_from_distance(dist: float, cutoff: float = 3.5) -> float:
    return max(0.0, min(1.0, (cutoff - dist) / 1.0))


def bond_type(atom_a: str, atom_b: str) -> str:
    if atom_a in BACKBONE_ATOMS and atom_b in BACKBONE_ATOMS:
        return "backbone"
    if atom_a in BACKBONE_ATOMS or atom_b in BACKBONE_ATOMS:
        return "mixed"
    return "sidechain"


def parse_excel(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb["1. IgV"]

    strand_by_col: dict[int, str] = {}
    for c in range(4, 17):
        v = ws.cell(6, c).value
        if v is not None and str(v).strip():
            strand_by_col[c] = str(v).strip()

    domain = DOMAIN
    title = ws.cell(1, 9).value
    if title and str(title).strip():
        domain = str(title).strip()

    records: list[dict] = []
    for row in ws.iter_rows(min_row=1, max_row=40, max_col=20):
        for cell in row:
            if cell.value is None:
                continue
            m = RES_PAT.match(str(cell.value).strip())
            if not m:
                continue

            aa = m.group(1)
            pdb_pos = int(m.group(2))
            ig_raw = ws.cell(cell.row, cell.column + 21).value
            if not isinstance(ig_raw, int):
                continue
            ig = int(ig_raw)

            segment = strand_by_col.get(cell.column) or segment_from_igstrand(ig)
            address = f"{get_column_letter(cell.column)}{cell.row}"

            records.append(
                {
                    "aa": aa,
                    "pdb_pos": pdb_pos,
                    "igstrand": ig,
                    "segment": segment,
                    "domain": domain,
                    "spreadsheet_cell": address,
                    "spreadsheet_row": cell.row,
                    "spreadsheet_col": cell.column,
                }
            )

    records.sort(key=lambda r: r["igstrand"])
    return records


def parse_interaction_html(path: Path) -> list[dict]:
    """Parse unique undirected H-bond atom pairs from iCn3D all-interaction HTML."""
    html = path.read_text(encoding="utf-8", errors="replace")
    bonds: list[dict] = []
    seen: set[tuple] = set()

    for dist_s, resid_a, resid_b in HBOND_ROW_PAT.findall(html):
        ma, mb = ATOM_PAT.search(resid_a), ATOM_PAT.search(resid_b)
        if not ma or not mb:
            continue
        aa1, u1, at1 = AA3_TO_1.get(ma.group(1), "?"), int(ma.group(2)), ma.group(3)
        aa2, u2, at2 = AA3_TO_1.get(mb.group(1), "?"), int(mb.group(2)), mb.group(3)
        key = tuple(sorted([(u1, at1), (u2, at2)]))
        if key in seen:
            continue
        seen.add(key)
        dist = float(dist_s)
        bonds.append(
            {
                "uniprot_a": u1,
                "aa_a": aa1,
                "atom_a": at1,
                "uniprot_b": u2,
                "aa_b": aa2,
                "atom_b": at2,
                "distance": dist,
                "weight": round(weight_from_distance(dist), 4),
                "type": bond_type(at1, at2),
            }
        )
    return bonds


def detect_uniprot_offset(pdb_aa: dict[int, str], bonds: list[dict]) -> int:
    """Find UniProt→PDB offset by maximizing AA identity matches."""
    html_aa: dict[int, str] = {}
    for b in bonds:
        html_aa[b["uniprot_a"]] = b["aa_a"]
        html_aa[b["uniprot_b"]] = b["aa_b"]

    best = (0, 21)  # (matches, offset); 21 is the known 1CD8/P01732 shift
    for offset in range(-40, 41):
        matches = sum(
            1
            for pdb_pos, aa in pdb_aa.items()
            if html_aa.get(pdb_pos + offset) == aa
        )
        if matches > best[0]:
            best = (matches, offset)
    if best[0] == 0:
        raise ValueError("Could not map UniProt positions from interaction HTML to Excel PDB numbers")
    return best[1]


def empty_interaction() -> dict:
    return {
        "hbond_density": 0,
        "hbond_count": 0,
        "backbone_count": 0,
        "sidechain_count": 0,
        "mixed_count": 0,
        "weight_total": 0.0,
        "bonds": [],
    }


def attach_hbonds(
    residues: list[dict],
    bonds: list[dict],
    offset: int,
) -> dict:
    by_pdb = {
        r["layers"]["1_identity"]["pdb"]["pos"]: r for r in residues
    }

    for rec in residues:
        ident = rec["layers"]["1_identity"]
        pdb_pos = ident["pdb"]["pos"]
        ident["uniprot"] = {"id": UNIPROT_ID, "pos": pdb_pos + offset}
        rec["layers"]["4_interaction"] = empty_interaction()

    edges: list[dict] = []
    for b in bonds:
        pdb_a = b["uniprot_a"] - offset
        pdb_b = b["uniprot_b"] - offset
        on_a = pdb_a in by_pdb
        on_b = pdb_b in by_pdb
        if not on_a and not on_b:
            continue

        # Prefer map residues; skip if AA disagrees with Excel (safety)
        if on_a and by_pdb[pdb_a]["layers"]["1_identity"]["residue"] != b["aa_a"]:
            continue
        if on_b and by_pdb[pdb_b]["layers"]["1_identity"]["residue"] != b["aa_b"]:
            continue

        edge = {
            "a": {
                "pdb": pdb_a,
                "uniprot": b["uniprot_a"],
                "atom": b["atom_a"],
                "aa": b["aa_a"],
                "on_map": on_a,
            },
            "b": {
                "pdb": pdb_b,
                "uniprot": b["uniprot_b"],
                "atom": b["atom_b"],
                "aa": b["aa_b"],
                "on_map": on_b,
            },
            "distance": b["distance"],
            "weight": b["weight"],
            "type": b["type"],
        }
        if on_a and on_b:
            edges.append(edge)

        for self_pdb, partner_pdb, self_atom, partner_atom, self_u, partner_u in (
            (pdb_a, pdb_b, b["atom_a"], b["atom_b"], b["uniprot_a"], b["uniprot_b"]),
            (pdb_b, pdb_a, b["atom_b"], b["atom_a"], b["uniprot_b"], b["uniprot_a"]),
        ):
            if self_pdb not in by_pdb:
                continue
            layer = by_pdb[self_pdb]["layers"]["4_interaction"]
            layer["bonds"].append(
                {
                    "partner_pdb": partner_pdb if partner_pdb in by_pdb else None,
                    "partner_uniprot": partner_u,
                    "partner_on_map": partner_pdb in by_pdb,
                    "self_atom": self_atom,
                    "partner_atom": partner_atom,
                    "distance": b["distance"],
                    "weight": b["weight"],
                    "type": b["type"],
                    "sep": abs(self_u - partner_u),
                }
            )

    for rec in residues:
        layer = rec["layers"]["4_interaction"]
        bonds_list = layer["bonds"]
        layer["hbond_count"] = len(bonds_list)
        layer["hbond_density"] = len(bonds_list)
        layer["backbone_count"] = sum(1 for x in bonds_list if x["type"] == "backbone")
        layer["sidechain_count"] = sum(1 for x in bonds_list if x["type"] == "sidechain")
        layer["mixed_count"] = sum(1 for x in bonds_list if x["type"] == "mixed")
        layer["weight_total"] = round(sum(x["weight"] for x in bonds_list), 4)
        bonds_list.sort(key=lambda x: (x["distance"], x["partner_uniprot"] or 0))

    return {
        "source": "icn3d_allinteraction",
        "uniprot_offset": offset,  # uniprot_pos = pdb_pos + offset
        "unique_pairs_parsed": len(bonds),
        "edges_on_map": len(edges),
        "edges": edges,
    }


def build(excel_path: Path, interaction_path: Path | None) -> dict:
    records = parse_excel(excel_path)

    residues = []
    for seq_num, rec in enumerate(records, start=1):
        ig = rec["igstrand"]
        layer1 = {
            "seq_num": seq_num,
            "residue": rec["aa"],
            "chain": CHAIN,
            "domain": rec["domain"],
            "segment": rec["segment"],
            "template_position": {
                "row": rec["spreadsheet_row"],
                "col": rec["spreadsheet_col"],
            },
            "spreadsheet_cell": rec["spreadsheet_cell"],
            "spreadsheet_position": {
                "row": rec["spreadsheet_row"],
                "col": rec["spreadsheet_col"],
                "address": rec["spreadsheet_cell"],
            },
            "foldnum": {
                "scheme": "IgStrand",
                "code": ig,
                "anchor": is_anchor(ig),
            },
            "uniprot": {"id": UNIPROT_ID, "pos": None},
            "pdb": {"id": PDB_ID, "chain": CHAIN, "pos": rec["pdb_pos"]},
        }

        residues.append(
            {
                "id": f"{PDB_ID}:{CHAIN}:{rec['pdb_pos']}",
                "layers": {
                    "1_identity": layer1,
                    **empty_layers(),
                },
            }
        )

    hbonds_block = None
    if interaction_path is not None:
        raw_bonds = parse_interaction_html(interaction_path)
        pdb_aa = {
            r["layers"]["1_identity"]["pdb"]["pos"]: r["layers"]["1_identity"]["residue"]
            for r in residues
        }
        offset = detect_uniprot_offset(pdb_aa, raw_bonds)
        hbonds_block = attach_hbonds(residues, raw_bonds, offset)

    dens_vals = [
        r["layers"]["4_interaction"].get("hbond_density", 0)
        for r in residues
        if r["layers"]["4_interaction"]
    ]

    doc = {
        "schema_version": "0.2",
        "fold": "IgV",
        "protein": {
            "id": UNIPROT_ID,
            "pdb": PDB_ID,
            "name": "CD8A",
            "domain": DOMAIN,
        },
        "template": {
            "id": "excel:1CD8_ig_diagram",
            "numbering_scheme": "IgStrand",
            "source": str(excel_path),
        },
        "sources": {
            "excel": str(excel_path),
            "pdb": {"id": PDB_ID, "chain": CHAIN},
            "uniprot": {"id": UNIPROT_ID},
            "interaction_html": str(interaction_path) if interaction_path else None,
        },
        "layer_defs": {
            "1_identity": {
                "description": "Addressing & identity from Excel (+ UniProt pos from interaction map)",
                "fields": [
                    "seq_num",
                    "residue",
                    "chain",
                    "domain",
                    "segment",
                    "template_position",
                    "spreadsheet_cell",
                    "foldnum",
                    "uniprot",
                    "pdb",
                ],
            },
            "2_intrinsic": {"description": "From amino-acid type alone"},
            "3_structural": {"description": "From 3D model"},
            "4_interaction": {
                "description": "H-bond density & partners from iCn3D interaction table",
                "fields": [
                    "hbond_density",
                    "hbond_count",
                    "backbone_count",
                    "sidechain_count",
                    "mixed_count",
                    "weight_total",
                    "bonds",
                ],
            },
            "5_comparative": {"description": "MSA / mutations"},
        },
        "counts": {
            "residues": len(residues),
            "anchors": sum(
                1 for r in residues if r["layers"]["1_identity"]["foldnum"]["anchor"]
            ),
            "segments": sorted(
                {
                    r["layers"]["1_identity"]["segment"]
                    for r in residues
                    if r["layers"]["1_identity"]["segment"]
                }
            ),
            "hbond_edges_on_map": hbonds_block["edges_on_map"] if hbonds_block else 0,
            "residues_with_hbonds": sum(1 for d in dens_vals if d > 0),
            "hbond_density_max": max(dens_vals) if dens_vals else 0,
        },
        "residues": residues,
    }
    if hbonds_block is not None:
        doc["hbonds"] = hbonds_block
    return doc


def main() -> None:
    p = argparse.ArgumentParser(
        description="Build CD8A layered JSON from Excel + iCn3D interaction HTML"
    )
    p.add_argument(
        "excel",
        nargs="?",
        type=Path,
        default=EXCEL_DEFAULT,
        help=f"Path to 1CD8_ig_diagram.xlsx (default: {EXCEL_DEFAULT})",
    )
    p.add_argument(
        "--interactions",
        type=Path,
        default=INTERACTION_DEFAULT,
        help=f"iCn3D all-interaction HTML (default: {INTERACTION_DEFAULT})",
    )
    p.add_argument(
        "--skip-hbonds",
        action="store_true",
        help="Build Layer 1 only (ignore interaction HTML)",
    )
    p.add_argument(
        "-o",
        "--output",
        type=Path,
        default=OUT_DEFAULT,
        help=f"Output JSON path (default: {OUT_DEFAULT})",
    )
    args = p.parse_args()

    excel = args.excel.expanduser().resolve()
    interaction = None if args.skip_hbonds else args.interactions.expanduser().resolve()
    if interaction is not None and not interaction.exists():
        raise SystemExit(f"Interaction HTML not found: {interaction}")

    doc = build(excel, interaction)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(doc, indent=2))
    print(f"wrote {args.output}")
    print(
        f"residues={doc['counts']['residues']} "
        f"anchors={doc['counts']['anchors']} "
        f"hb_edges={doc['counts']['hbond_edges_on_map']} "
        f"with_hb={doc['counts']['residues_with_hbonds']} "
        f"density_max={doc['counts']['hbond_density_max']}"
    )
    if "hbonds" in doc:
        print(f"uniprot_offset=+{doc['hbonds']['uniprot_offset']} (uniprot = pdb + offset)")


if __name__ == "__main__":
    main()
